"""
Management command to load water quality data from Excel file.
"""
from django.core.management.base import BaseCommand
from django.utils import timezone
from datetime import datetime
import openpyxl
import os
from monitoring.models import WaterBody, WaterQualityMeasurement


class Command(BaseCommand):
    help = 'Loads water bodies and measurements from Excel file'

    def handle(self, *args, **options):
        # Path to Excel file
        excel_path = os.path.join(
            os.path.dirname(__file__),
            'water_quality_data.xlsx'
        )
        
        if not os.path.exists(excel_path):
            self.stdout.write(
                self.style.ERROR(f'Excel file not found: {excel_path}')
            )
            return
        
        self.stdout.write(f'Loading data from: {excel_path}')
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        
        # Load Water Bodies
        self.stdout.write('\nLoading water bodies...')
        ws_wb = wb['WaterBodies']
        water_bodies = []
        
        for row in ws_wb.iter_rows(min_row=2, values_only=True):
            name, wb_type, lat, lon, description, regulatory_body = row
            
            wb_obj, created = WaterBody.objects.get_or_create(
                name=name,
                defaults={
                    'water_body_type': wb_type,
                    'latitude': lat,
                    'longitude': lon,
                    'description': description,
                    'regulatory_body': regulatory_body
                }
            )
            water_bodies.append(wb_obj)
            
            if created:
                self.stdout.write(
                    self.style.SUCCESS(f'  ✓ Created: {name}')
                )
            else:
                self.stdout.write(f'  - Already exists: {name}')
        
        # Load Measurements
        self.stdout.write('\nLoading measurements...')
        ws_meas = wb['Measurements']
        measurements_created = 0
        measurements_with_alerts = 0
        
        for row in ws_meas.iter_rows(min_row=2, values_only=True):
            (water_body_name, date_str, ph, dissolved_oxygen, temperature,
             turbidity, nitrates, phosphates, ecoli_count, measured_by,
             sample_lat, sample_lon, notes) = row
            
            # Find water body
            try:
                water_body = WaterBody.objects.get(name=water_body_name)
            except WaterBody.DoesNotExist:
                self.stdout.write(
                    self.style.WARNING(f'  ⚠️  Water body not found: {water_body_name}')
                )
                continue
            
            # Parse date
            if isinstance(date_str, str):
                measured_at = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                measured_at = date_str
            
            # Make timezone aware
            measured_at = timezone.make_aware(measured_at)
            
            # Create measurement
            measurement = WaterQualityMeasurement.objects.create(
                water_body=water_body,
                measured_at=measured_at,
                ph=ph,
                dissolved_oxygen=dissolved_oxygen,
                temperature=temperature,
                turbidity=turbidity,
                nitrates=nitrates,
                phosphates=phosphates,
                ecoli_count=ecoli_count,
                measured_by=measured_by,
                sample_latitude=sample_lat,
                sample_longitude=sample_lon,
                notes=notes
            )
            measurements_created += 1
            
            # Show alert status
            if measurement.has_alerts:
                measurements_with_alerts += 1
                alerts = measurement.get_alerts()
                self.stdout.write(
                    self.style.WARNING(
                        f'  ⚠️  {water_body_name} - {measured_at.strftime("%Y-%m-%d")} - '
                        f'{len(alerts)} alert(s): {", ".join(alerts[:2])}'
                    )
                )
        
        # Display summary
        self.stdout.write(
            self.style.SUCCESS(
                f'\n✓ Successfully loaded {measurements_created} measurements for {len(water_bodies)} water bodies'
            )
        )
        self.stdout.write(
            self.style.SUCCESS(
                f'✓ Measurements with EPA alerts: {measurements_with_alerts}'
            )
        )
