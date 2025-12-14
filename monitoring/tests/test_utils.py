"""
Tests for AquaWatch utilities.
"""
import pytest
from io import BytesIO
from openpyxl import load_workbook
from monitoring.utils import export_data_to_excel
from monitoring.models import WaterQualityMeasurement


@pytest.mark.django_db
class TestExportUtility:
    """Tests for data export utility."""
    
    def test_export_creates_excel_file(self, measurement):
        """Test export creates valid Excel file."""
        queryset = WaterQualityMeasurement.objects.all()
        response = export_data_to_excel(queryset)
        
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'attachment' in response['Content-Disposition']
    
    def test_export_includes_measurement_data(self, measurement):
        """Test exported file includes measurement data."""
        queryset = WaterQualityMeasurement.objects.all()
        response = export_data_to_excel(queryset)
        
        # Load workbook from response
        wb = load_workbook(BytesIO(response.content))
        ws = wb['Water Quality Data']
        
        # Check headers
        assert ws['A1'].value == 'Water Body'
        assert ws['G1'].value == 'pH'
        
        # Check data row exists
        assert ws['A2'].value == 'Test Lake'
        assert ws['G2'].value == 7.5
    
    def test_export_includes_summary_sheet(self, measurement):
        """Test export includes summary sheet."""
        queryset = WaterQualityMeasurement.objects.all()
        response = export_data_to_excel(queryset)
        
        wb = load_workbook(BytesIO(response.content))
        assert 'Summary' in wb.sheetnames
        
        summary_ws = wb['Summary']
        assert 'Export Summary' in summary_ws['A1'].value
    
    def test_export_includes_epa_standards(self, measurement):
        """Test export includes EPA standards reference."""
        queryset = WaterQualityMeasurement.objects.all()
        response = export_data_to_excel(queryset)
        
        wb = load_workbook(BytesIO(response.content))
        assert 'EPA Standards' in wb.sheetnames
        
        standards_ws = wb['EPA Standards']
        assert 'EPA' in standards_ws['A1'].value
    
    def test_export_highlights_non_compliant(self, water_body):
        """Test export highlights non-compliant measurements."""
        # Create non-compliant measurement
        WaterQualityMeasurement.objects.create(
            water_body=water_body,
            measured_at=water_body.monitoring_start_date,
            ph=5.0,  # Below EPA standard
            dissolved_oxygen=8.0,
            temperature=20.0,
            turbidity=3.0,
            nitrates=5.0,
            phosphates=0.05,
            ecoli_count=50,
            measured_by="Test"
        )
        
        queryset = WaterQualityMeasurement.objects.all()
        response = export_data_to_excel(queryset)
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb['Water Quality Data']
        
        # Check compliance status column
        assert ws['N2'].value == 'NON-COMPLIANT'
