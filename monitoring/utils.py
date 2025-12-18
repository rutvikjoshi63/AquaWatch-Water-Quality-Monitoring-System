from openpyxl import Workbook
from datetime import datetime
from django.http import HttpResponse

def export_data_to_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Water Quality Data"
    
    headers = [
        'Water Body', 'Water Body Type', 'Latitude', 'Longitude',
        'Measurement Date/Time','pH', 'Dissolved Oxygen (mg/L)', 
        'Temperature (°C)','Notes'
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        
    row_num = 2
    for measurement in queryset.select_related('water_body'):
        row_data = [
            measurement.water_body.name,
            measurement.water_body.get_water_body_type_display(),
            measurement.water_body.latitude,
            measurement.water_body.longitude,
            measurement.measured_at.strftime('%Y-%m-%d %H:%M:%S'),
            measurement.ph,
            measurement.dissolved_oxygen,
            measurement.temperature,
            measurement.notes
        ]
        
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col, value=value)
        
        row_num += 1
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'aquawatch_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response
