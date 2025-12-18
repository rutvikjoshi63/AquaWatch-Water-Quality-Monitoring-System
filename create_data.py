from django.core.management.base import BaseCommand
import os
import openpyxl
from monitoring.models import WaterBody, WaterQualityMeasurement
from datetime import datetime

class Command(BaseCommand):
    def handle(self, *args, **options):
        excel_path = os.path.join(
            os.path.dirname(__file__),
            'water_quality_data.xlsx'
        )
        wb = openpyxl.load_workbook(excel_path)
        ws_wb = wb['WaterBodies']
        for row in ws_wb.iter_rows(min_row=2, values_only=True):
            name, wb_type, lat, lon, description, regulatory_body = row
            
            WaterBody.objects.get_or_create(
                name=name,
                defaults={
                    'water_body_type': wb_type,
                    'latitude': lat,
                    'longitude': lon,
                    'description': description,
                    'regulatory_body': regulatory_body
                }
            )
        ws_meas = wb['Measurements']
        
        for row in ws_meas.iter_rows(min_row=2, values_only=True):
            (water_body_name, date_str, ph, dissolved_oxygen, temperature,
            sample_lat, sample_lon, notes) = row
            
            try:
                water_body = WaterBody.objects.get(name=water_body_name)
            except WaterBody.DoesNotExist:
                continue
            
            if isinstance(date_str, str):
                measured_at = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                measured_at = date_str
            
            
            WaterQualityMeasurement.objects.create(
                water_body=water_body,
                measured_at=measured_at,
                ph=ph,
                dissolved_oxygen=dissolved_oxygen,
                temperature=temperature,
                notes=notes
            )
